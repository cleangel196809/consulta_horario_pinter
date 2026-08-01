"""
Servicio de importación de los archivos Excel institucionales:
  - PLANEACIÓN (hojas: PLANEACION, DOCENTES, REFLEJOS, CERRADOS)
  - INSCRITOS_POR_CICLO (hoja única con estudiantes matriculados)

Este módulo NO asume nombres de archivo fijos: solo depende de los
encabezados de columna, que se detectan de forma flexible (mayúsculas/
minúsculas, espacios extra, tildes).
"""
import re
import unicodedata
from datetime import time as dtime
from io import BytesIO
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy.dialects.postgresql import insert as pg_insert

from .. import models
from .usuarios_auto import sincronizar_usuario_docente, sincronizar_usuario_estudiante

SEDES_CONOCIDAS = [
    "SUR", "NORTE", "CENTRO", "CALLE 73", "CALLE 72", "CHAPINERO",
    "OCCIDENTE", "SOACHA", "KENNEDY", "VIRTUAL",
]


def _norm(s) -> str:
    """Normaliza un encabezado: sin tildes, minúsculas, sin espacios extra."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s)
    return s


def _header_map(header_row) -> dict:
    """Devuelve {nombre_normalizado: indice_columna} para una fila de encabezado."""
    mapping = {}
    for idx, val in enumerate(header_row):
        key = _norm(val)
        if key:
            mapping[key] = idx
    return mapping


def _get(row, mapping, *names, default=None):
    for name in names:
        idx = mapping.get(_norm(name))
        if idx is not None and idx < len(row):
            value = row[idx]
            if value is not None and str(value).strip() != "":
                return value
    return default


def _to_int(value) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _to_time(value) -> Optional[dtime]:
    if value is None or value == "":
        return None
    if isinstance(value, dtime):
        return value
    s = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _extraer_sede(nombre_salon: Optional[str], modalidad: Optional[str]) -> Optional[str]:
    """La sede real a veces viene como sufijo de NOMBRE_SALON ('SALON 403-SEDE SUR')
    y en otras filas queda registrada en la columna MODALIDAD por costumbre de captura
    del archivo institucional. Se intenta extraer de ambas fuentes."""
    for text in (nombre_salon, modalidad):
        if not text:
            continue
        t = _norm(text)
        m = re.search(r"SEDE\s+([A-Z0-9 ]+)", t)
        if m:
            return m.group(1).strip().title()
        for sede in SEDES_CONOCIDAS:
            if sede in t:
                return sede.title()
    return None


def _find_header_row(ws, must_contain, max_scan: int = 20):
    """Busca la fila de encabezado escaneando las primeras `max_scan` filas.

    `must_contain` es una lista donde cada elemento puede ser:
      - un nombre de columna (str), o
      - un grupo de nombres alternativos/sinónimos (list[str]), del cual
        basta que aparezca UNO en el encabezado (útil para hojas como
        REFLEJOS que renombran ASIGNATURA/GRUPO como "... VIGENTE").
    """
    grupos = [([m] if isinstance(m, str) else list(m)) for m in must_contain]
    grupos_norm = [[_norm(alias) for alias in grupo] for grupo in grupos]

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        norm_row = [_norm(v) for v in row]
        if all(any(alias == v for v in norm_row for alias in grupo) for grupo in grupos_norm):
            return i + 1, row  # openpyxl es 1-indexado
    raise ValueError(
        f"No se encontró la fila de encabezado esperada (se buscaban las columnas: {must_contain})"
    )


def import_docentes_desde_planeacion(db: Session, wb, carga_id: int) -> tuple[int, int]:
    """Importa/actualiza la hoja DOCENTES (NIT_CC, NOMBRE_COMPLETO, CORREO, FACULTAD, SEDE).

    Al final de la importación, sincroniza automáticamente el usuario de
    login de cada docente con correo institucional @pi.edu.co (ver
    `services/usuarios_auto.py`).
    """
    if "DOCENTES" not in wb.sheetnames:
        return 0, 0
    ws = wb["DOCENTES"]
    header_row_num, header_row = _find_header_row(ws, ["NIT_CC", "NOMBRE_COMPLETO"])
    mapping = _header_map(header_row)

    procesadas, errores = 0, 0
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        try:
            cedula = _to_int(_get(row, mapping, "NIT_CC"))
            nombre = _to_str(_get(row, mapping, "NOMBRE_COMPLETO"))
            if not cedula or not nombre:
                continue
            correo = _to_str(_get(row, mapping, "CORREO INSTITUCIONAL", "CORREO_INSTITUCIONAL"))
            facultad = _to_str(_get(row, mapping, "FACULTAD"))
            sede = _to_str(_get(row, mapping, "SEDE"))
            if facultad and facultad.upper() in ("#N/A", "N/A"):
                facultad = None
            if sede and sede.upper() in ("#N/A", "N/A"):
                sede = None

            stmt = pg_insert(models.Docente).values(
                cedula=cedula,
                nombre_completo=nombre,
                correo_institucional=correo,
                facultad=facultad,
                sede=sede,
            ).on_conflict_do_update(
                index_elements=["cedula"],
                set_=dict(
                    nombre_completo=nombre,
                    correo_institucional=correo,
                    facultad=facultad,
                    sede=sede,
                ),
            )
            db.execute(stmt)
            procesadas += 1

            # Sincronización de usuario de login del docente. Se aísla en su
            # propio try/except para que un problema al crear/actualizar el
            # usuario NUNCA haga fallar la importación del docente en sí.
            try:
                sincronizar_usuario_docente(db, cedula, nombre, correo)
            except Exception:
                pass
        except Exception:
            errores += 1
    db.commit()
    return procesadas, errores


def _extraer_filas_horario(wb, hoja: str) -> tuple[list[dict], int]:
    """Parsea una hoja de horarios (PLANEACION, REFLEJOS o CERRADOS) y
    devuelve una lista de diccionarios con los campos ya normalizados
    (listos para instanciar `models.Horario`, sin `periodo`/`carga_id`/
    `origen_hoja`, que se agregan al momento de insertar), junto con el
    conteo de filas que no se pudieron parsear.

    Esta función NO toca la base de datos: se usa tanto para insertar
    (`import_horarios_desde_planeacion`) como para previsualizar duplicados
    (`detectar_duplicados_planeacion`) sin efectos secundarios.
    """
    if hoja not in wb.sheetnames:
        return [], 0
    ws = wb[hoja]
    header_row_num, header_row = _find_header_row(
        ws, [["ASIGNATURA", "ASIGNATURA VIGENTE"], ["GRUPO", "GRUPO ASIGNATURA VIGENTE"]]
    )
    mapping = _header_map(header_row)

    filas = []
    errores = 0
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        try:
            grupo = _to_str(_get(row, mapping, "GRUPO", "GRUPO ASIGNATURA VIGENTE"))
            if not grupo:
                continue

            docente_cedula = _to_int(_get(row, mapping, "DOCUMENTO_DOCENTE"))
            if docente_cedula is not None and docente_cedula <= 0:
                # Algunas filas traen 0 como marcador de "docente sin asignar";
                # se trata como ausencia de docente para no violar la llave foránea.
                docente_cedula = None
            nombre_docente = _to_str(_get(row, mapping, "NOMBRE_DOCENTE"))
            correo_docente = _to_str(_get(row, mapping, "CORREO INSTITUCIONAL", "CORREO_INSTITUCIONAL"))

            nombre_salon = _to_str(_get(row, mapping, "NOMBRE_SALON"))
            modalidad = _to_str(_get(row, mapping, "MODALIDAD"))
            sede = _extraer_sede(nombre_salon, modalidad)

            fila = dict(
                llave=_to_str(_get(row, mapping, "LLAVE")),
                codigo_asignatura=_to_str(_get(row, mapping, "CODIGO", "CODIGO ASIGNATURA VIGENTE")),
                facultad=_to_str(_get(row, mapping, "FACULTAD")),
                programa=_to_str(_get(row, mapping, "PROGRAMA", "PROGRAMA ASIGNATURA VIGENTE")),
                plan=_to_str(_get(row, mapping, "PLAN")),
                asignatura=_to_str(_get(row, mapping, "ASIGNATURA", "ASIGNATURA VIGENTE")),
                ciclo=_to_str(_get(row, mapping, "CICLO", "CICLO ASIGNATURA VIGENTE")),
                creditos=_to_str(_get(row, mapping, "CREDITOS", "NUMERO DE CREDITOS ASIGNATURA VIGENTE")),
                grupo=grupo,
                codigo_moodle=_to_str(_get(row, mapping, "CODIGO MOODLE")),
                codigo_teams=_to_str(_get(row, mapping, "CODIGO TEAMS")),
                enlace_teams=_to_str(_get(row, mapping, "ENLACE TEAMS")),
                estado=_to_str(_get(row, mapping, "ESTADO")),
                modalidad=modalidad,
                jornada=_to_str(_get(row, mapping, "JORNADA")),
                capacidad=_to_int(_get(row, mapping, "CAPACIDAD")),
                dia=_to_str(_get(row, mapping, "DIA")),
                hora_inicio=_to_time(_get(row, mapping, "H_INICIO")),
                hora_fin=_to_time(_get(row, mapping, "H_FIN")),
                nombre_salon=nombre_salon,
                sede=sede,
                docente_cedula=docente_cedula,
                nombre_docente=nombre_docente,
                correo_docente=correo_docente,
                observaciones=_to_str(_get(row, mapping, "OBSERVACIONES")),
            )
            filas.append(fila)
        except Exception:
            errores += 1
    return filas, errores


def _clave_duplicado(fila: dict) -> tuple:
    """Define cuándo dos filas de horario son 'la misma clase': deben
    coincidir EXACTAMENTE en docente + día + hora inicio/fin + salón + grupo
    + asignatura. Se incluye `dia` a propósito para NO confundir con un
    falso positivo una misma clase que legítimamente se repite en días
    distintos de la semana (p. ej. lunes y miércoles) — esas filas tienen
    `dia` diferente y por lo tanto una clave distinta, así que no se
    consideran duplicadas ni se eliminan.
    """
    return (
        fila.get("docente_cedula"),
        fila.get("dia"),
        fila.get("hora_inicio"),
        fila.get("hora_fin"),
        fila.get("nombre_salon"),
        fila.get("grupo"),
        fila.get("asignatura"),
    )


def detectar_duplicados_planeacion(wb, periodo: str) -> dict:
    """Analiza (SIN insertar nada en la base de datos) el archivo de
    PLANEACIÓN y devuelve un resumen de filas duplicadas, entendiendo por
    duplicado dos filas que coinciden exactamente en docente_cedula + dia +
    hora_inicio + hora_fin + nombre_salon + grupo + asignatura, dentro del
    periodo que se va a cargar. Se analizan en conjunto las hojas
    PLANEACION, REFLEJOS y CERRADOS (las tres terminan insertándose en la
    misma tabla `horarios`).
    """
    grupos_por_clave: dict[tuple, list[dict]] = {}
    for hoja in ("PLANEACION", "REFLEJOS", "CERRADOS"):
        filas, _errores = _extraer_filas_horario(wb, hoja)
        for fila in filas:
            clave = _clave_duplicado(fila)
            grupos_por_clave.setdefault(clave, []).append({**fila, "origen_hoja": hoja})

    grupos_duplicados = []
    total_duplicados = 0
    for clave, filas in grupos_por_clave.items():
        if len(filas) <= 1:
            continue
        total_duplicados += len(filas) - 1  # filas "de más" que sobrarían tras deduplicar
        primera = filas[0]
        grupos_duplicados.append({
            "docente_cedula": clave[0],
            "nombre_docente": primera.get("nombre_docente"),
            "dia": clave[1],
            "hora_inicio": str(clave[2]) if clave[2] else None,
            "hora_fin": str(clave[3]) if clave[3] else None,
            "nombre_salon": clave[4],
            "grupo": clave[5],
            "asignatura": clave[6],
            "veces_repetido": len(filas),
            "hojas": [f["origen_hoja"] for f in filas],
        })

    grupos_duplicados.sort(key=lambda g: g["veces_repetido"], reverse=True)

    return {
        "periodo": periodo,
        "duplicados_encontrados": total_duplicados,
        "grupos_duplicados": grupos_duplicados,
    }


def import_horarios_desde_planeacion(
    db: Session,
    wb,
    periodo: str,
    carga_id: int,
    hoja: str = "PLANEACION",
    eliminar_duplicados: bool = False,
    claves_vistas: Optional[set] = None,
) -> tuple[int, int]:
    """Importa la hoja de horarios (PLANEACION, REFLEJOS o CERRADOS).

    Si `eliminar_duplicados=True`, se conserva solo la primera fila de cada
    combinación duplicada (según `_clave_duplicado`) y se descartan las
    demás antes de insertar. `claves_vistas` permite compartir el conjunto
    de claves ya vistas entre varias llamadas (una por hoja), para que la
    deduplicación funcione también entre PLANEACION/REFLEJOS/CERRADOS y no
    solo dentro de una misma hoja.
    """
    filas, errores = _extraer_filas_horario(wb, hoja)
    if claves_vistas is None:
        claves_vistas = set()

    # Aseguramos que los docentes referenciados existan (aunque no vengan en la
    # hoja DOCENTES) para no romper la llave foránea.
    docentes_vistos: dict[int, bool] = {}

    procesadas = 0
    for fila in filas:
        try:
            if eliminar_duplicados:
                clave = _clave_duplicado(fila)
                if clave in claves_vistas:
                    continue
                claves_vistas.add(clave)

            docente_cedula = fila.get("docente_cedula")
            if docente_cedula is not None and docente_cedula not in docentes_vistos:
                stmt = pg_insert(models.Docente).values(
                    cedula=docente_cedula,
                    nombre_completo=fila.get("nombre_docente") or f"Docente {docente_cedula}",
                    correo_institucional=fila.get("correo_docente"),
                ).on_conflict_do_nothing(index_elements=["cedula"])
                db.execute(stmt)
                docentes_vistos[docente_cedula] = True

            horario = models.Horario(
                periodo=periodo,
                origen_hoja=hoja,
                carga_id=carga_id,
                **fila,
            )
            db.add(horario)
            procesadas += 1
        except Exception:
            errores += 1
    db.commit()
    return procesadas, errores


def importar_planeacion(
    db: Session, file_bytes: bytes, periodo: str, carga_id: int, eliminar_duplicados: bool = False
) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    total_ok, total_err = 0, 0

    ok, err = import_docentes_desde_planeacion(db, wb, carga_id)
    total_ok += ok
    total_err += err

    # `claves_vistas` se comparte entre las tres hojas para que la
    # deduplicación (cuando se solicita) funcione de forma consistente con
    # `detectar_duplicados_planeacion`, que también las analiza en conjunto.
    claves_vistas: set = set()
    for hoja in ("PLANEACION", "REFLEJOS", "CERRADOS"):
        ok, err = import_horarios_desde_planeacion(
            db, wb, periodo, carga_id, hoja=hoja,
            eliminar_duplicados=eliminar_duplicados, claves_vistas=claves_vistas,
        )
        total_ok += ok
        total_err += err

    return {"filas_procesadas": total_ok, "filas_error": total_err}


def importar_inscritos(db: Session, file_bytes: bytes, periodo: str, carga_id: int) -> dict:
    """Importa el archivo INSCRITOS_POR_CICLO (hoja única, con filas de título
    antes del encabezado real). Al finalizar, sincroniza automáticamente el
    usuario de login de cada estudiante con correo @pi.edu.co (ver
    `services/usuarios_auto.py`)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row_num, header_row = _find_header_row(ws, ["IDENTIFICACION", "GRUPO", "ASIGNATURA"])
    mapping = _header_map(header_row)

    procesadas, errores = 0, 0
    estudiantes_vistos = set()
    # Info mínima de cada estudiante visto en esta carga, para poder
    # sincronizar su usuario de login al final sin tener que releer el Excel.
    estudiantes_info: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        try:
            cedula = _to_str(_get(row, mapping, "IDENTIFICACION"))
            if not cedula:
                continue

            if cedula not in estudiantes_vistos:
                nombres = _to_str(_get(row, mapping, "NOMBRES"))
                apellidos = _to_str(_get(row, mapping, "APELLIDOS"))
                email = _to_str(_get(row, mapping, "EMAIL"))
                correo_aula_virtual = _to_str(_get(row, mapping, "CORREO AULA VIRTUAL"))

                stmt = pg_insert(models.Estudiante).values(
                    cedula=cedula,
                    tipo=_to_str(_get(row, mapping, "TIPO")),
                    nombres=nombres,
                    apellidos=apellidos,
                    correo_aula_virtual=correo_aula_virtual,
                    email=email,
                    celular=_to_str(_get(row, mapping, "CELULAR")),
                    telefono=_to_str(_get(row, mapping, "TELEFONO")),
                    ciclo_ingreso=_to_str(_get(row, mapping, "CICLO_INGRESO")),
                ).on_conflict_do_update(
                    index_elements=["cedula"],
                    set_=dict(
                        tipo=_to_str(_get(row, mapping, "TIPO")),
                        nombres=nombres,
                        apellidos=apellidos,
                        correo_aula_virtual=correo_aula_virtual,
                        email=email,
                        celular=_to_str(_get(row, mapping, "CELULAR")),
                        telefono=_to_str(_get(row, mapping, "TELEFONO")),
                        ciclo_ingreso=_to_str(_get(row, mapping, "CICLO_INGRESO")),
                    ),
                )
                db.execute(stmt)
                estudiantes_vistos.add(cedula)
                estudiantes_info[cedula] = {
                    "nombre_completo": " ".join(p for p in (nombres, apellidos) if p) or None,
                    "email": email,
                    "correo_aula_virtual": correo_aula_virtual,
                }

            inscripcion_stmt = pg_insert(models.Inscripcion).values(
                estudiante_cedula=cedula,
                periodo=periodo,
                ciclo_ingreso=_to_str(_get(row, mapping, "CICLO_INGRESO")),
                cod_plan=_to_str(_get(row, mapping, "COD_PLAN")),
                nom_plan=_to_str(_get(row, mapping, "NOM_PLAN")),
                cod_asignatura=_to_str(_get(row, mapping, "COD_ASIGNATURA")),
                asignatura=_to_str(_get(row, mapping, "ASIGNATURA")),
                ciclo=_to_str(_get(row, mapping, "CICLO")),
                creditos=_to_str(_get(row, mapping, "CREDITOS")),
                grupo=_to_str(_get(row, mapping, "GRUPO")),
                jornada=_to_str(_get(row, mapping, "JORNADA")),
                estado=_to_str(_get(row, mapping, "ESTADO")),
                sede=_to_str(_get(row, mapping, "SEDE")),
                identificador=_to_str(_get(row, mapping, "IDENTIFICADOR")),
                flg_virtual=_to_str(_get(row, mapping, "FLG_VIRTUAL")),
                nombre_facultad=_to_str(_get(row, mapping, "NOMBRE_FACULTAD")),
                semilla=_to_str(_get(row, mapping, "SEMILLA")),
                carga_id=carga_id,
            ).on_conflict_do_update(
                index_elements=["estudiante_cedula", "periodo", "cod_asignatura", "grupo"],
                set_=dict(
                    estado=_to_str(_get(row, mapping, "ESTADO")),
                    jornada=_to_str(_get(row, mapping, "JORNADA")),
                    sede=_to_str(_get(row, mapping, "SEDE")),
                    carga_id=carga_id,
                ),
            )
            db.execute(inscripcion_stmt)
            procesadas += 1
        except Exception:
            errores += 1

    # Sincronización de usuarios de login de estudiantes (aislada en su
    # propio try/except por estudiante para no afectar el resultado de la
    # importación de inscripciones si algo falla al crear el usuario).
    for cedula, info in estudiantes_info.items():
        try:
            sincronizar_usuario_estudiante(
                db, cedula, info["nombre_completo"], info["email"], info["correo_aula_virtual"]
            )
        except Exception:
            pass

    db.commit()
    return {"filas_procesadas": procesadas, "filas_error": errores}
